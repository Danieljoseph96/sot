from io import BytesIO
from decimal import Decimal

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import LocalityWise, UserReg


class ExportColumnTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="tester", password="secret123")
        self.client.login(username="tester", password="secret123")
        UserReg.objects.create(
            name="Alice",
            locality="Chennai",
            state="Tamil Nadu",
            language="Tamil",
            total_amount="125.50",
            balance_amount="10.00",
        )
        UserReg.objects.create(
            name="Bala",
            locality="Madurai",
            state="Tamil Nadu",
            language="Tamil",
            total_amount="200.00",
            balance_amount="0.00",
        )
        UserReg.objects.create(
            name="Carol",
            locality="Bengaluru",
            state="Karnataka",
            language="Kannada",
            total_amount="300.00",
            balance_amount="25.00",
        )
        UserReg.objects.filter(name="Alice").update(acc_room="xx yy zz")
        UserReg.objects.filter(name="Bala").update(acc_room="aa bb")
        UserReg.objects.filter(name="Carol").update(acc_room="xx cc")

    def test_export_page_uses_selected_columns(self):
        response = self.client.get(
            reverse("export_page"),
            {"columns": ["name", "locality", "total_amount"]},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["selected_export_columns"],
            ["name", "locality", "total_amount"],
        )
        self.assertEqual(
            response.context["export_headers"],
            ["Name", "Locality", "Total Amount"],
        )
        self.assertEqual(
            response.context["export_preview_rows"][0],
            ["Alice", "Chennai", "125.50"],
        )

    def test_export_xlsx_downloads_selected_columns(self):
        response = self.client.get(
            reverse("export_xlsx"),
            {"columns": ["name", "state"]},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(
            [worksheet.cell(row=1, column=1).value, worksheet.cell(row=1, column=2).value],
            ["Name", "State"],
        )
        self.assertEqual(
            [worksheet.cell(row=2, column=1).value, worksheet.cell(row=2, column=2).value],
            ["Alice", "Tamil Nadu"],
        )

    def test_export_page_applies_multi_select_filters(self):
        response = self.client.get(
            reverse("export_page"),
            {
                "columns": ["name", "locality", "language"],
                "filter_field": ["state", "language"],
                "filter_value": ["Tamil Nadu", "Tamil"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["export_total_count"], 2)
        self.assertEqual(
            response.context["selected_export_filter_rows"],
            [
                {"field": "state", "value": "Tamil Nadu"},
                {"field": "language", "value": "Tamil"},
            ],
        )
        self.assertEqual(
            response.context["export_preview_rows"],
            [
                ["Alice", "Chennai", "Tamil"],
                ["Bala", "Madurai", "Tamil"],
            ],
        )

    def test_export_pdf_uses_selected_columns_and_filters(self):
        response = self.client.get(
            reverse("export_pdf"),
            {
                "columns": ["name", "state"],
                "filter_field": ["state"],
                "filter_value": ["Karnataka"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(b"Carol", response.content)
        self.assertIn(b"Karnataka", response.content)
        self.assertNotIn(b"Alice", response.content)

    def test_export_page_still_supports_legacy_filter_query_params(self):
        response = self.client.get(
            reverse("export_page"),
            {
                "columns": ["name", "locality"],
                "state": ["Tamil Nadu"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["selected_export_filter_rows"],
            [{"field": "state", "value": "Tamil Nadu"}],
        )
        self.assertEqual(response.context["export_total_count"], 2)

    def test_export_page_acc_room_filter_uses_first_value_before_space(self):
        response = self.client.get(
            reverse("export_page"),
            {
                "columns": ["name", "acc_room"],
                "filter_field": ["acc_room"],
                "filter_value": ["xx"],
            },
        )

        self.assertEqual(response.status_code, 200)
        acc_room_filter = next(
            option for option in response.context["export_filter_options"] if option["name"] == "acc_room"
        )
        self.assertEqual(acc_room_filter["values"], ["aa", "xx"])
        self.assertEqual(
            response.context["export_preview_rows"],
            [
                ["Alice", "xx yy zz"],
                ["Carol", "xx cc"],
            ],
        )


class SearchSummaryTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="tester", password="secret123")
        self.client.login(username="tester", password="secret123")
        UserReg.objects.create(
            name="Anu",
            bs="Sister",
            age=19,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="100.00",
            balance_amount="0.00",
        )
        UserReg.objects.create(
            name="Balu",
            bs="Brother",
            age=22,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="150.00",
            balance_amount="25.00",
        )
        UserReg.objects.create(
            name="Cathy",
            bs="Sister",
            age=17,
            locality="Madurai",
            state="Tamil Nadu",
            total_amount="200.00",
            balance_amount="10.00",
        )

    def test_search_locality_shows_summary_and_paid_status_when_payment_method_not_pending(self):
        LocalityWise.objects.create(
            locality="Chennai",
            state="Tamil Nadu",
            persons_count=2,
            total_paid="250.00",
            total_balance="25.00",
            payment_method="cash",
        )

        response = self.client.get(reverse("search"), {"q": "Chennai", "columns": ["name", "locality"]})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["result_count"], 2)
        self.assertEqual(
            response.context["search_summary"],
            {
                "locality": "Chennai",
                "matched_localities": ["Chennai"],
                "total_people": 2,
                "total_sister": 1,
                "total_brother": 1,
                "total_teenager": 1,
                "total_amount_sum": "250",
                "balance_amount_sum": "25",
                "current_status": "Paid",
                "has_locality_summary": True,
                "status_badge_class": "text-bg-success",
            },
        )

    def test_search_locality_without_summary_record_marks_did_not_pay(self):
        response = self.client.get(reverse("search"), {"q": "Madurai", "columns": ["name", "locality"]})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["result_count"], 1)
        self.assertEqual(response.context["search_summary"]["current_status"], "Did not pay")
        self.assertFalse(response.context["search_summary"]["has_locality_summary"])
        self.assertEqual(response.context["search_summary"]["status_badge_class"], "text-bg-danger")

    def test_search_locality_with_pending_payment_method_marks_did_not_pay(self):
        LocalityWise.objects.create(
            locality="Chennai",
            state="Tamil Nadu",
            persons_count=2,
            total_paid="0.00",
            total_balance="25.00",
            payment_method="pending",
        )

        response = self.client.get(reverse("search"), {"q": "Chennai", "columns": ["name", "locality"]})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["search_summary"]["current_status"], "Did not pay")
        self.assertEqual(response.context["search_summary"]["status_badge_class"], "text-bg-danger")

    def test_search_page_lists_paid_localities_filtered_by_search_query(self):
        LocalityWise.objects.create(
            locality="Chennai",
            state="Tamil Nadu",
            persons_count=2,
            total_paid="250.00",
            total_balance="0.00",
            payment_method="cash",
        )
        LocalityWise.objects.create(
            locality="Madurai",
            state="Tamil Nadu",
            persons_count=1,
            total_paid="100.00",
            total_balance="0.00",
            payment_method="upi",
        )
        LocalityWise.objects.create(
            locality="Salem",
            state="Tamil Nadu",
            persons_count=1,
            total_paid="0.00",
            total_balance="10.00",
            payment_method="pending",
        )

        response = self.client.get(reverse("search"), {"q": "Chennai", "columns": ["name", "locality"]})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["paid_localities"],
            [
                {
                    "locality": "Chennai",
                    "state": "Tamil Nadu",
                    "total_paid": Decimal("250.00"),
                    "payment_method": "cash",
                }
            ],
        )


class LocalityRegisterTests(TestCase):
    def setUp(self):
        UserReg.objects.create(
            name="Anu",
            bs="Sister",
            age=19,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="100.00",
            balance_amount="10.00",
        )
        UserReg.objects.create(
            name="Balu",
            bs="Brother",
            age=22,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="200.00",
            balance_amount="20.00",
        )
        UserReg.objects.create(
            name="Cathy",
            bs="Sister",
            age=17,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="50.00",
            balance_amount="5.00",
        )
        UserReg.objects.create(
            name="Deepa",
            bs="Sister",
            age=10,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="25.00",
            balance_amount="5.00",
        )
        UserReg.objects.create(
            name="Esh",
            bs="Brother",
            age=3,
            locality="Chennai",
            state="Tamil Nadu",
            total_amount="10.00",
            balance_amount="0.00",
        )

    def test_locality_register_page_renders_form(self):
        response = self.client.get(reverse("reg"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Locality register form")
        self.assertIn("locality_frm", response.context)
        self.assertContains(response, "Select locality")

    def test_locality_register_summary_endpoint_returns_userreg_aggregate(self):
        response = self.client.get(reverse("locality_register_summary"), {"locality": "Chennai"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {
                "locality": "Chennai",
                "state": "Tamil Nadu",
                "persons_count": 5,
                "brother_count": 2,
                "sister_count": 3,
                "children_count": 1,
                "infant_count": 1,
                "teenager_count": 2,
                "total_amount_sum": "385",
                "balance_amount_sum": "40",
            },
        )

    def test_locality_register_post_saves_with_auto_loaded_summary_values(self):
        response = self.client.post(
            reverse("reg"),
            {
                "locality": "Chennai",
                "state": "",
                "persons_count": 0,
                "total_paid": "100.00",
                "total_balance": "0.00",
                "payment_method": "cash",
                "remarks": "Test entry",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("reg"))
        locality_summary = LocalityWise.objects.get(locality="Chennai")
        self.assertEqual(locality_summary.state, "Tamil Nadu")
        self.assertEqual(locality_summary.persons_count, 5)
        self.assertEqual(locality_summary.total_balance, 40)
        self.assertEqual(locality_summary.total_paid, 100)

    def test_paid_locality_hidden_from_dropdown_and_listed_separately(self):
        LocalityWise.objects.create(
            locality="Chennai",
            state="Tamil Nadu",
            persons_count=5,
            total_paid="385.00",
            total_balance="40.00",
            payment_method="cash",
        )

        response = self.client.get(reverse("reg"))

        locality_choices = [choice[0] for choice in response.context["locality_frm"].fields["locality"].choices]
        self.assertNotIn("Chennai", locality_choices)
        self.assertEqual(
            response.context["paid_localities"],
            [
                {
                    "locality": "Chennai",
                    "state": "Tamil Nadu",
                    "total_paid": Decimal("385.00"),
                    "payment_method": "cash",
                }
            ],
        )


class HomeAndUserRegPageTests(TestCase):
    def setUp(self):
        UserReg.objects.create(
            name="Anu",
            bs="Sister",
            locality="Chennai",
            state="Tamil Nadu",
            language="Tamil",
            total_amount="100.00",
            balance_amount="10.00",
        )
        UserReg.objects.create(
            name="Balu",
            bs="Brother",
            locality="Madurai",
            state="Tamil Nadu",
            language="Tamil",
            total_amount="120.00",
            balance_amount="0.00",
        )
        LocalityWise.objects.create(
            locality="Chennai",
            state="Tamil Nadu",
            persons_count=1,
            total_paid="100.00",
            total_balance="0.00",
            payment_method="cash",
        )
        LocalityWise.objects.create(
            locality="Madurai",
            state="Tamil Nadu",
            persons_count=1,
            total_paid="0.00",
            total_balance="120.00",
            payment_method="pending",
        )

    def test_home_page_contains_multiple_chart_datasets(self):
        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["payment_chart_labels"], ["Cash", "Pending"])
        self.assertEqual(response.context["payment_chart_values"], [1, 1])
        self.assertEqual(response.context["registration_chart_labels"], ["Chennai", "Madurai"])
        self.assertEqual(response.context["registration_chart_values"], [1, 1])
        self.assertEqual(response.context["locality_chart_labels"], ["Chennai", "Madurai"])
        self.assertEqual(response.context["locality_chart_values"], [1, 1])

    def test_userreg_page_filters_by_dropdown_values(self):
        response = self.client.get(reverse("userreg_list"), {"bs": "Sister", "state": "Tamil Nadu"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["userreg_count"], 1)
        self.assertEqual(
            response.context["selected_userreg_filters"],
            {"bs": "Sister", "state": "Tamil Nadu"},
        )
        self.assertEqual(
            response.context["userreg_rows"],
            [["1", "Anu", "Sister", "Chennai", "Tamil Nadu", "Tamil", "100.00", "10.00"]],
        )
