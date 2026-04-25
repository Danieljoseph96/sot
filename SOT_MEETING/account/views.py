from io import BytesIO
import re
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time
from urllib.parse import urlencode
from itertools import zip_longest

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import models as dj_models
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, legal, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import models as app_models
from . import forms as app_forms
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors
from django.db import connection
from django.shortcuts import render
from io import BytesIO
from django.http import HttpResponse
from django.db import connection
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import landscape, legal
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch




PENDING_IMPORT_ROWS_SESSION_KEY = "pending_import_rows"
PENDING_IMPORT_SHEET_SESSION_KEY = "pending_import_sheet_name"
PENDING_IMPORT_TABLE_NAME_SESSION_KEY = "pending_import_table_name"
PENDING_IMPORT_OPTIONS_SESSION_KEY = "pending_import_options"
PENDING_IMPORT_HEADER_ROW_SESSION_KEY = "pending_import_header_row"
PENDING_IMPORT_DB_NAME_SESSION_KEY = "pending_import_db_name"
PENDING_IMPORT_HEADER_WARNINGS_SESSION_KEY = "pending_import_header_warnings"

DATE_INPUT_FORMATS = [
    "%d.%m.%y",
	"%Y-%m-%d",
	"%Y/%m/%d",
	"%d/%m/%Y",
	"%m/%d/%Y",
	"%d-%m-%Y",
]

TIME_INPUT_FORMATS = ["%H:%M", "%H:%M:%S"]

EXCLUDED_USERREG_IMPORT_FIELDS = {"sl_no", "created_at", "updated_at"}




def get_report_rows():
    rows = list(
        app_models.UserReg.objects.order_by("sl_no").values_list("name", "total_amount","acc_on_1_2")
    )
    return [(name, float(amount or 0), acc_on_1_2) for name, amount, acc_on_1_2 in rows]


USERREG_FILTER_FIELDS = ("bs", "locality", "state", "language")


def get_dashboard_chart_data():
    locality_rows = list(
        app_models.UserReg.objects.exclude(locality__isnull=True)
        .exclude(locality="")
        .values("locality")
        .annotate(total=dj_models.Count("sl_no"))
        .order_by("-total", "locality")[:10]
    )
    payment_rows = list(
        app_models.LocalityWise.objects.exclude(payment_method__isnull=True)
        .exclude(payment_method="")
        .values("payment_method")
        .annotate(total=dj_models.Count("locality"))
        .order_by("payment_method")
    )
    registration_rows = list(
        app_models.LocalityWise.objects.exclude(locality__isnull=True)
        .exclude(locality="")
        .values("locality")
        .annotate(total=dj_models.Sum("persons_count"))
        .order_by("-total", "locality")[:10]
    )
    return {
        "locality_chart_labels": [row["locality"] for row in locality_rows],
        "locality_chart_values": [row["total"] for row in locality_rows],
        "payment_chart_labels": [row["payment_method"].title() for row in payment_rows],
        "payment_chart_values": [row["total"] for row in payment_rows],
        "registration_chart_labels": [row["locality"] for row in registration_rows],
        "registration_chart_values": [int(row["total"] or 0) for row in registration_rows],
    }


def get_userreg_export_fields():
    return list(app_models.UserReg._meta.fields)


def format_field_label(field_name):
    return field_name.replace("_", " ").title()


def get_userreg_export_column_options():
    return [
        {
            "name": field.name,
            "label": format_field_label(field.name),
        }
        for field in get_userreg_export_fields()
    ]


def get_valid_export_column_names():
    return {field.name for field in get_userreg_export_fields()}


def resolve_selected_export_columns(request):
    valid_column_names = get_valid_export_column_names()
    requested_columns = [
        column_name
        for column_name in request.GET.getlist("columns")
        if column_name in valid_column_names
    ]
    if requested_columns:
        return requested_columns
    return [field.name for field in get_userreg_export_fields()]


EXPORT_FILTER_FIELDS = (
    "locality",
    "state",
    "language",
    "acc",
    "acc_on_1_2",
    "transport",
)


VALID_ACC_CODES = {
    "ANR", "AMR", "CFR", "SS", "ARCADE",
    "LUX ADB", "LUX P", "LUX S", "LUX C",
    "LUX AA", "LUX AB", "LUX AC",
    "LUX N", "LUX V",
    "GGH",
    "D1", "D2", "D3", "D4",
    "VC", "V DB",
    "SELF",
}

def normalize_export_filter_value(field_name, raw_value):
    text = str(raw_value).strip()

    if field_name == "acc_on_1_2":
        words = text.upper().split()

        if not words:
            return ""

        # Handle 2-word codes
        if len(words) >= 2:
            two_word = f"{words[0]} {words[1]}"
            if two_word in VALID_ACC_CODES:
                return two_word

        # Handle 1-word codes
        one_word = words[0]
        if one_word in VALID_ACC_CODES:
            return one_word

        return ""

    return text

def get_export_filter_options(selected_filters=None):
    options = []
    base_queryset = app_models.UserReg.objects.order_by("sl_no")
    for field_name in EXPORT_FILTER_FIELDS:
        raw_values = (
            base_queryset.exclude(**{f"{field_name}__isnull": True})
            .exclude(**{field_name: ""})
            .order_by(field_name)
            .values_list(field_name, flat=True)
            .distinct()
        )
        normalized_values = []
        seen_values = set()
        for raw_value in raw_values:
            normalized_value = normalize_export_filter_value(field_name, raw_value)
            if not normalized_value or normalized_value in seen_values:
                continue
            seen_values.add(normalized_value)
            normalized_values.append(normalized_value)
        options.append(
            {
                "name": field_name,
                "label": format_field_label(field_name),
                "values": normalized_values,
            }
        )
    return options


def get_export_filter_value_map():
    return {
        option["name"]: set(option["values"])
        for option in get_export_filter_options()
    }


def build_legacy_export_filter_rows(request):
    rows = []
    for field_name in EXPORT_FILTER_FIELDS:
        selected_values = [
            value.strip()
            for value in request.GET.getlist(field_name)
            if value.strip()
        ]
        for value in selected_values:
            rows.append({"field": field_name, "value": value})
    return rows


def resolve_export_filter_rows(request):
    valid_fields = set(EXPORT_FILTER_FIELDS)
    valid_value_map = get_export_filter_value_map()
    requested_fields = request.GET.getlist("filter_field")
    requested_values = request.GET.getlist("filter_value")
    rows = []

    for field_name, selected_value in zip_longest(requested_fields, requested_values, fillvalue=""):
        field_name = (field_name or "").strip()
        selected_value = (selected_value or "").strip()
        if not field_name or not selected_value:
            continue
        if field_name not in valid_fields:
            continue
        if selected_value not in valid_value_map.get(field_name, set()):
            continue
        rows.append({"field": field_name, "value": selected_value})

    if rows:
        return rows
    return build_legacy_export_filter_rows(request)


def apply_export_filters(queryset, filter_rows):
    filtered_queryset = queryset

    for filter_row in filter_rows:
        field_name = filter_row["field"]
        value = filter_row["value"].strip()

        # Special filter for acc_on_1_2
        if field_name == "acc_on_1_2":

            filtered_queryset = filtered_queryset.filter(
                Q(acc_on_1_2__iexact=value) |
                Q(acc_on_1_2__istartswith=value + " ")
            )
            continue

        filtered_queryset = filtered_queryset.filter(**{field_name: value})

    return filtered_queryset


def build_export_query_string(selected_export_columns, filter_rows):
    params = [("columns", column) for column in selected_export_columns]
    for filter_row in filter_rows:
        params.append(("filter_field", filter_row["field"]))
        params.append(("filter_value", filter_row["value"]))
    return urlencode(params, doseq=True)


def populate_export_context(context, request, queryset=None):
    selected_export_columns = resolve_selected_export_columns(request)
    selected_filter_rows = resolve_export_filter_rows(request)
    filtered_queryset = queryset or apply_export_filters(
        app_models.UserReg.objects.order_by("sl_no"),
        selected_filter_rows,
    )
    export_headers, export_rows = get_userreg_export_dataset(selected_export_columns, filtered_queryset)
    context["export_column_options"] = get_userreg_export_column_options()
    context["export_filter_options"] = get_export_filter_options()
    context["selected_export_filter_rows"] = selected_filter_rows or [{"field": "", "value": ""}]
    context["selected_export_columns"] = selected_export_columns
    context["export_headers"] = export_headers
    context["export_preview_rows"] = export_rows[:20]
    context["export_total_count"] = len(export_rows)
    context["export_preview_count"] = len(context["export_preview_rows"])
    context["export_query_string"] = build_export_query_string(selected_export_columns, selected_filter_rows)
    context["export_active_filter_count"] = len(selected_filter_rows)
    return context


def format_export_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value)


def get_userreg_export_dataset(column_names, queryset=None):
    export_queryset = queryset or app_models.UserReg.objects.order_by("sl_no")
    headers = [format_field_label(column_name) for column_name in column_names]
    rows = []
    for record in export_queryset.values(*column_names):
        rows.append([format_export_cell_value(record.get(column_name)) for column_name in column_names])
    return headers, rows


def get_export_pdf_pagesize(column_count):
    if column_count >= 8:
        return landscape(legal)
    return landscape(letter)


def get_export_pdf_col_widths(headers, rows, available_width):
    column_count = max(len(headers), 1)
    if column_count == 1:
        return [available_width]

    raw_widths = []
    for index, header in enumerate(headers):
        sample_values = [header]
        sample_values.extend(row[index] for row in rows[:25] if index < len(row))
        longest = max((len(str(value)) for value in sample_values), default=8)
        raw_widths.append(max(1, min(longest, 30)))

    total_raw_width = sum(raw_widths) or column_count
    scaled_widths = [(available_width * width) / total_raw_width for width in raw_widths]
    min_width = min(90, available_width / column_count)
    adjusted_widths = [max(min_width, width) for width in scaled_widths]
    adjusted_total = sum(adjusted_widths)

    if adjusted_total > available_width:
        scale = available_width / adjusted_total
        adjusted_widths = [width * scale for width in adjusted_widths]

    return adjusted_widths


def build_export_pdf_table_data(headers, rows, styles):
    header_style = styles["Heading6"]
    header_style.fontName = "Helvetica-Bold"
    header_style.fontSize = 8
    header_style.leading = 10
    header_style.textColor = colors.white
    header_style.wordWrap = "CJK"

    cell_style = styles["BodyText"]
    cell_style.fontName = "Helvetica"
    cell_style.fontSize = 7
    cell_style.leading = 9
    cell_style.wordWrap = "CJK"

    table_data = [[Paragraph(str(header), header_style) for header in headers]]
    if rows:
        for row in rows:
            table_data.append([Paragraph(str(cell), cell_style) for cell in row])
    else:
        table_data.append(
            [Paragraph("No rows available for the selected filters.", cell_style)]
            + [Paragraph("", cell_style) for _ in range(max(len(headers) - 1, 0))]
        )
    return table_data


def get_site_urls(request):
	return [
		{"label": "Home", "path": reverse("home")},
		{"label": "Register", "path": reverse("reg")},
		{"label": "UserReg", "path": reverse("userreg_list")},
		{"label": "Search", "path": reverse("search")},
		{"label": "Export page", "path": reverse("export_page")},
		{"label": "Import page", "path": reverse("import_page")},
		{"label": "Download XLSX", "path": reverse("export_xlsx")},
		{"label": "Download PDF", "path": reverse("export_pdf")},
	]


def build_common_context(request):
    report_rows = get_report_rows()
    dashboard_chart_data = get_dashboard_chart_data()
    report_data = {
        "title": "SOT Meeting Report",
        "chart_labels": [row[0] for row in report_rows],
        "chart_values": [row[1] for row in report_rows],
        "chart_acc_on_1_2s": [row[2] for row in report_rows],

    }
    return {
        "chart_labels": report_data["chart_labels"],
        "chart_values": report_data["chart_values"],
        "chart_acc_on_1_2s": report_data["chart_acc_on_1_2s"],
        "locality_chart_labels": dashboard_chart_data["locality_chart_labels"],
        "locality_chart_values": dashboard_chart_data["locality_chart_values"],
        "payment_chart_labels": dashboard_chart_data["payment_chart_labels"],
        "payment_chart_values": dashboard_chart_data["payment_chart_values"],
        "registration_chart_labels": dashboard_chart_data["registration_chart_labels"],
        "registration_chart_values": dashboard_chart_data["registration_chart_values"],
        "report_rows": report_rows,
        "site_urls": get_site_urls(request),
    }


def login_view(request):
    if request.user.is_authenticated:
        return redirect("home")

    next_url = request.GET.get("next") or request.POST.get("next") or reverse("home")
    context = {"next": next_url}

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is None:
            context["login_error"] = "Invalid username or password."
            context["username"] = username
            return render(request, "account/login.html", context, status=400)

        login(request, user)
        return redirect(next_url)

    return render(request, "account/login.html", context)


@login_required(login_url="login")
def logout_view(request):
    if request.method == "POST":
        logout(request)
        return redirect("login")
    return redirect("home")


#@login_required(login_url="login")
def home(request):
	context = build_common_context(request)
	return render(request, "account/home.html", context)


def userreg_list(request):
    queryset = app_models.UserReg.objects.order_by("sl_no")
    selected_filters = resolve_userreg_filters(request)
    filtered_queryset = apply_userreg_filters(queryset, selected_filters)
    selected_columns = [
        "sl_no",
        "name",
        "bs",
        "locality",
        "state",
        "language",
        "total_amount",
        "balance_amount",
    ]
    table_headers, table_rows = get_userreg_export_dataset(selected_columns, filtered_queryset)

    context = build_common_context(request)
    context.update(
        {
            "userreg_filter_options": get_userreg_filter_options(selected_filters),
            "selected_userreg_filters": selected_filters,
            "userreg_headers": table_headers,
            "userreg_rows": table_rows,
            "userreg_count": len(table_rows),
        }
    )
    return render(request, "account/userreg_list.html", context)


def normalize_header_name(header):
	if header is None:
		return ""
	text = str(header).strip().lower()
	text = re.sub(r"[^a-z0-9]+", "_", text)
	return text.strip("_")


def get_userreg_import_fields():
	result = {}
	for field in app_models.UserReg._meta.fields:
		if field.name in EXCLUDED_USERREG_IMPORT_FIELDS:
			continue
		result[field.name] = field
	return result


def get_required_userreg_fields(field_map):
	required = set()
	for field in field_map.values():
		if field.primary_key or isinstance(field, dj_models.AutoField):
			continue
		if field.null or field.blank or field.has_default():
			continue
		required.add(field.name)
	return required


def validate_import_headers(header_row):
	field_map = get_userreg_import_fields()
	required_fields = get_required_userreg_fields(field_map)

	index_to_field = {}
	unknown_headers = []
	normalized_headers = []

	for index, header in enumerate(header_row):
		normalized = normalize_header_name(header)
		normalized_headers.append(normalized)
		if normalized in field_map:
			index_to_field[index] = normalized
		elif normalized:
			unknown_headers.append(str(header))

	mapped_fields = set(index_to_field.values())
	missing_required = sorted(required_fields - mapped_fields)

	warnings = []
	if unknown_headers:
		warnings.append(
			"Ignored unknown header column(s): " + ", ".join(unknown_headers)
		)
	if missing_required:
		warnings.append(
			"Missing required UserReg field column(s): "
			+ ", ".join(missing_required)
			+ "."
		)

	return {
		"index_to_field": index_to_field,
		"field_map": field_map,
		"missing_required": missing_required,
		"warnings": warnings,
		"normalized_headers": normalized_headers,
	}


def parse_date_value(raw_value):
	if isinstance(raw_value, datetime):
		return raw_value.date()
	if isinstance(raw_value, date):
		return raw_value

	text = str(raw_value).strip()
	if not text:
		raise ValueError("empty date")

	for fmt in DATE_INPUT_FORMATS:
		try:
			return datetime.strptime(text, fmt).date()
		except ValueError:
			continue

	try:
		return date.fromisoformat(text)
	except ValueError as exc:
		raise ValueError("invalid date format") from exc


def parse_time_value(raw_value):
	if isinstance(raw_value, datetime):
		return raw_value.time().replace(microsecond=0)
	if isinstance(raw_value, time):
		return raw_value.replace(microsecond=0)

	text = str(raw_value).strip()
	if not text:
		raise ValueError("empty time")

	for fmt in TIME_INPUT_FORMATS:
		try:
			return datetime.strptime(text, fmt).time()
		except ValueError:
			continue

	try:
		return time.fromisoformat(text)
	except ValueError as exc:
		raise ValueError("invalid time format") from exc


def convert_cell_to_field_value(raw_value, field):
	if raw_value is None:
		if field.null:
			return None
		if isinstance(field, (dj_models.CharField, dj_models.TextField)):
			return ""
		raise ValueError("required value is empty")

	text = str(raw_value).strip()
	if text == "":
		if field.null:
			return None
		if isinstance(field, (dj_models.CharField, dj_models.TextField)):
			return ""
		raise ValueError("required value is empty")

	if isinstance(field, dj_models.DateField):
		return parse_date_value(raw_value)

	if isinstance(field, dj_models.TimeField):
		return parse_time_value(raw_value)

	if isinstance(field, dj_models.PositiveIntegerField):
		number = int(float(text))
		if number < 0:
			raise ValueError("must be zero or greater")
		return number

	if isinstance(field, dj_models.IntegerField):
		return int(float(text))

	if isinstance(field, dj_models.DecimalField):
		try:
			return Decimal(text)
		except InvalidOperation as exc:
			raise ValueError("invalid decimal value") from exc

	if isinstance(field, dj_models.BooleanField):
		lowered = text.lower()
		if lowered in ("1", "true", "yes", "y"):
			return True
		if lowered in ("0", "false", "no", "n"):
			return False
		raise ValueError("invalid boolean value")

	return text


#@login_required(login_url="login")
def export_page(request):
    context = build_common_context(request)
    populate_export_context(context, request)
    return render(request, "account/export.html", context)


def populate_import_context(context, request):
    pending_rows = request.session.get(PENDING_IMPORT_ROWS_SESSION_KEY)
    pending_sheet_name = request.session.get(PENDING_IMPORT_SHEET_SESSION_KEY)
    pending_table_name = request.session.get(PENDING_IMPORT_TABLE_NAME_SESSION_KEY)
    pending_options = request.session.get(PENDING_IMPORT_OPTIONS_SESSION_KEY, {})
    pending_header_row = request.session.get(PENDING_IMPORT_HEADER_ROW_SESSION_KEY, [])
    pending_warnings = request.session.get(PENDING_IMPORT_HEADER_WARNINGS_SESSION_KEY, [])
    default_db_name = pending_table_name or pending_sheet_name or "userreg_import"
    import_db_name = request.session.get(PENDING_IMPORT_DB_NAME_SESSION_KEY, default_db_name)

    context["import_options"] = {
        "has_table_title_row": bool(pending_options.get("has_table_title_row", False)),
        "has_header_row": True,
    }
    if pending_rows:
        context["import_preview_rows"] = pending_rows
        context["import_sheet_name"] = pending_sheet_name
        context["import_table_name"] = pending_table_name
        context["import_header_row"] = pending_header_row
        context["import_db_name"] = import_db_name
        context["preview_count"] = len(pending_rows)
    if pending_warnings:
        context["import_warning"] = " ".join(pending_warnings)
    return context


def import_page(request):
    context = build_common_context(request)
    populate_import_context(context, request)
    return render(request, "account/import.html", context)


def normalize_cell_for_preview(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def parse_import_workbook(uploaded_file, options):
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise ValueError(
            "Could not read the Excel file. Please upload a valid .xlsx file."
        ) from exc

    worksheet = workbook.active
    sheet_name = worksheet.title
    parsed_rows = []
    table_name = ""
    start_index = 1
    max_columns = max(worksheet.max_column or 1, 1)
    header_row = [f"Column {column_index}" for column_index in range(1, max_columns + 1)]

    if options.get("has_table_title_row"):
        first_cell = worksheet.cell(row=1, column=1).value
        second_cell = worksheet.cell(row=1, column=2).value
        table_name = str(first_cell).strip() if first_cell is not None else ""
        if second_cell in (None, ""):
            start_index = 2
        else:
            table_name = ""

    if options.get("has_header_row", True):
        header_row = [
            normalize_cell_for_preview(
                worksheet.cell(row=start_index, column=column_index).value
            )
            or f"Column {column_index}"
            for column_index in range(1, max_columns + 1)
        ]
        start_index += 1

    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if index < start_index:
            continue

        row_values = list(row[:max_columns])
        if len(row_values) < max_columns:
            row_values.extend([None] * (max_columns - len(row_values)))

        if all(cell is None for cell in row_values):
            continue

        parsed_rows.append([index, [normalize_cell_for_preview(cell) for cell in row_values]])

    if not parsed_rows:
        raise ValueError("No importable rows were found in the Excel file.")

    header_validation = validate_import_headers(header_row)

    if not header_validation["index_to_field"]:
        raise ValueError(
            "None of the Excel headers match UserReg fields. "
            "Use headers like: name, locality, state, language, total_amount, balance_amount."
        )

    return {
        "sheet_name": sheet_name,
        "table_name": table_name,
        "header_row": header_row,
        "rows": parsed_rows,
        "header_warnings": header_validation["warnings"],
        "missing_required": header_validation["missing_required"],
    }


def get_import_options_from_request(request):
	return {
		"has_table_title_row": request.POST.get("has_table_title_row") == "on",
		"has_header_row": True,
	}


def attach_preview_context(context, pending_rows, pending_header_row, request):
	context["import_preview_rows"] = pending_rows
	context["import_header_row"] = pending_header_row
	context["preview_count"] = len(pending_rows)
	context["import_sheet_name"] = request.session.get(PENDING_IMPORT_SHEET_SESSION_KEY)
	context["import_table_name"] = request.session.get(PENDING_IMPORT_TABLE_NAME_SESSION_KEY)
	context["import_db_name"] = request.session.get(PENDING_IMPORT_DB_NAME_SESSION_KEY, "")


def build_search_summary(query, queryset):
    if not query:
        return None

    matched_localities = list(
        queryset.order_by()
        .exclude(locality__isnull=True)
        .exclude(locality="")
        .values_list("locality", flat=True)
        .distinct()
    )
    if not matched_localities:
        return None

    aggregates = queryset.aggregate(
        total_amount_sum=Sum("total_amount"),
        balance_amount_sum=Sum("balance_amount"),
    )
    sister_count = queryset.filter(bs__iexact="Sister").count()
    brother_count = queryset.filter(bs__iexact="Brother").count()
    teenager_count = queryset.filter(age__isnull=False, age__lt=20).count()

    locality_label = matched_localities[0] if len(matched_localities) == 1 else ", ".join(matched_localities)
    locality_summary = None
    if len(matched_localities) == 1:
        locality_summary = app_models.LocalityWise.objects.filter(locality__iexact=matched_localities[0]).first()

    if locality_summary and locality_summary.payment_method != "pending":
        current_status = "Paid"
        status_badge_class = "text-bg-success"
    else:
        current_status = "Did not pay"
        status_badge_class = "text-bg-danger"

    return {
        "locality": locality_label,
        "matched_localities": matched_localities,
        "total_people": queryset.count(),
        "total_sister": sister_count,
        "total_brother": brother_count,
        "total_teenager": teenager_count,
        "total_amount_sum": format_export_cell_value(aggregates.get("total_amount_sum") or Decimal("0")),
        "balance_amount_sum": format_export_cell_value(aggregates.get("balance_amount_sum") or Decimal("0")),
        "current_status": current_status,
        "has_locality_summary": bool(locality_summary),
        "status_badge_class": status_badge_class,
    }


def import_xlsx(request):
    if not request.user.is_authenticated:
        return redirect(f"{reverse('login')}?next={reverse('import_page')}")

    if request.method != "POST":
        return redirect("import_page")

    action = request.POST.get("action", "preview")
    context = build_common_context(request)
    populate_import_context(context, request)

    if action == "upload":
        pending_rows = request.session.get(PENDING_IMPORT_ROWS_SESSION_KEY)
        pending_header_row = request.session.get(PENDING_IMPORT_HEADER_ROW_SESSION_KEY, [])
        pending_options = request.session.get(PENDING_IMPORT_OPTIONS_SESSION_KEY, {})
        pending_warnings = request.session.get(PENDING_IMPORT_HEADER_WARNINGS_SESSION_KEY, [])

        if not pending_rows:
            context["import_error"] = "Load an Excel file first to preview the data before upload."
            return render(request, "account/import.html", context, status=400)

        header_validation = validate_import_headers(pending_header_row)
        if header_validation["missing_required"]:
            context["import_error"] = (
                "Upload blocked. Missing required column(s): "
                + ", ".join(header_validation["missing_required"])
            )
            attach_preview_context(context, pending_rows, pending_header_row, request)
            if pending_warnings:
                context["import_warning"] = " ".join(pending_warnings)
            return render(request, "account/import.html", context, status=400)

        mapping = header_validation["index_to_field"]
        field_map = header_validation["field_map"]
        required_fields = get_required_userreg_fields(field_map)

        created_objects = []
        invalid_rows = []
        duplicate_in_upload = 0
        duplicate_in_database = 0
        fingerprints = set()

        for source_row, row_values in pending_rows:
            payload = {}
            conversion_error = None

            for index, field_name in mapping.items():
                field = field_map[field_name]
                raw_cell = row_values[index] if index < len(row_values) else ""
                try:
                    payload[field_name] = convert_cell_to_field_value(raw_cell, field)
                except (TypeError, ValueError) as exc:
                    conversion_error = f"Row {source_row}, column {field_name}: {exc}"
                    break

            if conversion_error:
                invalid_rows.append(conversion_error)
                continue

            missing_required_values = [
                field_name
                for field_name in required_fields
                if field_name in payload and payload[field_name] in (None, "")
            ]
            if missing_required_values:
                invalid_rows.append(
                    f"Row {source_row}: required value(s) missing for {', '.join(missing_required_values)}"
                )
                continue

            fingerprint = tuple((key, payload.get(key)) for key in sorted(payload.keys()))
            if fingerprint in fingerprints:
                duplicate_in_upload += 1
                continue
            fingerprints.add(fingerprint)

            if app_models.UserReg.objects.filter(**payload).exists():
                duplicate_in_database += 1
                continue

            created_objects.append(app_models.UserReg(**payload))

        if not created_objects and not invalid_rows and (duplicate_in_upload or duplicate_in_database):
            context["import_warning"] = (
                "No new records were uploaded because all rows are duplicates "
                "(within upload or already in database)."
            )
            attach_preview_context(context, pending_rows, pending_header_row, request)
            return render(request, "account/import.html", context, status=200)

        if not created_objects and invalid_rows:
            context["import_error"] = "Upload failed. No valid rows remained after validation."
            context["import_warning"] = invalid_rows[0]
            attach_preview_context(context, pending_rows, pending_header_row, request)
            return render(request, "account/import.html", context, status=400)

        with transaction.atomic():
            app_models.UserReg.objects.bulk_create(created_objects)

        request.session.pop(PENDING_IMPORT_ROWS_SESSION_KEY, None)
        request.session.pop(PENDING_IMPORT_SHEET_SESSION_KEY, None)
        request.session.pop(PENDING_IMPORT_TABLE_NAME_SESSION_KEY, None)
        request.session.pop(PENDING_IMPORT_OPTIONS_SESSION_KEY, None)
        request.session.pop(PENDING_IMPORT_HEADER_ROW_SESSION_KEY, None)
        request.session.pop(PENDING_IMPORT_DB_NAME_SESSION_KEY, None)
        request.session.pop(PENDING_IMPORT_HEADER_WARNINGS_SESSION_KEY, None)

        context = build_common_context(request)
        populate_import_context(context, request)
        success_parts = [f"Uploaded {len(created_objects)} new UserReg row(s)."]
        warning_parts = []
        if duplicate_in_upload:
            warning_parts.append(
                f"Skipped {duplicate_in_upload} duplicate row(s) inside the uploaded file."
            )
        if duplicate_in_database:
            warning_parts.append(
                f"Skipped {duplicate_in_database} row(s) already present in database."
            )
        if invalid_rows:
            warning_parts.append(f"Skipped {len(invalid_rows)} invalid row(s).")

            # show first 10 errors
            error_preview = "<br>".join(invalid_rows[:10])

            if len(invalid_rows) > 10:
                error_preview += f"<br>... and {len(invalid_rows)-10} more errors."

            context["invalid_details"] = error_preview

        context["import_success"] = " ".join(success_parts)
        if warning_parts:
            context["import_warning"] = " ".join(warning_parts)
        context["import_options"] = {
            "has_table_title_row": bool(pending_options.get("has_table_title_row", False)),
            "has_header_row": True,
        }
        return render(request, "account/import.html", context)

    uploaded_file = request.FILES.get("xlsx_file")
    import_options = get_import_options_from_request(request)
    context["import_options"] = import_options

    if not uploaded_file:
        context["import_error"] = "Please choose an Excel file to load."
        return render(request, "account/import.html", context, status=400)

    if not uploaded_file.name.lower().endswith(".xlsx"):
        context["import_error"] = "Only .xlsx files are supported."
        return render(request, "account/import.html", context, status=400)

    try:
        preview_data = parse_import_workbook(uploaded_file, import_options)
    except ValueError as exc:
        context["import_error"] = str(exc)
        return render(request, "account/import.html", context, status=400)

    request.session[PENDING_IMPORT_ROWS_SESSION_KEY] = preview_data["rows"]
    request.session[PENDING_IMPORT_SHEET_SESSION_KEY] = preview_data["sheet_name"]
    request.session[PENDING_IMPORT_TABLE_NAME_SESSION_KEY] = preview_data["table_name"]
    request.session[PENDING_IMPORT_OPTIONS_SESSION_KEY] = import_options
    request.session[PENDING_IMPORT_HEADER_ROW_SESSION_KEY] = preview_data["header_row"]
    request.session[PENDING_IMPORT_DB_NAME_SESSION_KEY] = (
        preview_data["table_name"] or preview_data["sheet_name"] or "userreg_import"
    )
    request.session[PENDING_IMPORT_HEADER_WARNINGS_SESSION_KEY] = preview_data.get(
        "header_warnings", []
    )

    context["import_sheet_name"] = preview_data["sheet_name"]
    context["import_table_name"] = preview_data["table_name"]
    context["import_header_row"] = preview_data["header_row"]
    context["import_db_name"] = request.session.get(PENDING_IMPORT_DB_NAME_SESSION_KEY)
    context["import_preview_rows"] = preview_data["rows"]
    context["preview_count"] = len(preview_data["rows"])
    context["import_success"] = "Excel loaded successfully. Review the table, then click Upload to DB."

    if preview_data.get("header_warnings"):
        context["import_warning"] = " ".join(preview_data["header_warnings"])

    return render(request, "account/import.html", context)

#@login_required(login_url="login")
def search(request):

	query = request.GET.get("q", "").strip()
	queryset = app_models.UserReg.objects.order_by("sl_no")

	if query:
		locality_queryset = queryset.filter(locality__icontains=query)
		if locality_queryset.exists():
			queryset = locality_queryset
		else:
			search_filter = Q(name__icontains=query)
			search_filter |= Q(locality__icontains=query)
			search_filter |= Q(state__icontains=query)
			if query.replace(".", "", 1).isdigit():
				search_filter |= Q(total_amount=Decimal(query))
				search_filter |= Q(balance_amount=Decimal(query))
			queryset = queryset.filter(search_filter)

	selected_export_columns = resolve_selected_export_columns(request)
	search_headers, search_rows = get_userreg_export_dataset(selected_export_columns, queryset)
	search_summary = build_search_summary(query, queryset)

	context = build_common_context(request)
	context.update(
		{
			"search_query": query,
			"search_results": search_rows,
			"search_headers": search_headers,
			"result_count": len(search_rows),
            "export_column_options": get_userreg_export_column_options(),
            "selected_export_columns": selected_export_columns,
            "search_summary": search_summary,
            "paid_localities": get_search_paid_localities(query),
		}
	)
	return render(request, "search.html", context)


def export_xlsx(request):
    if not request.user.is_authenticated:
        return redirect(f"{reverse('login')}?next={reverse('export_xlsx')}")

    selected_export_columns = resolve_selected_export_columns(request)
    selected_filter_rows = resolve_export_filter_rows(request)
    filtered_queryset = apply_export_filters(
        app_models.UserReg.objects.order_by("sl_no"),
        selected_filter_rows,
    )
    export_headers, export_rows = get_userreg_export_dataset(selected_export_columns, filtered_queryset)
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "UserReg Data"
    worksheet.append(export_headers)

    for row in export_rows:
        worksheet.append(row)

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="sot-meeting-report.xlsx"'
    return response

def export_pdf(request):
    selected_export_columns = resolve_selected_export_columns(request)
    selected_filter_rows = resolve_export_filter_rows(request)

    locality_name = ""
    state_name = ""

    for row in selected_filter_rows:
        if row["field"] == "locality":
            locality_name = row["value"].strip()
            break

    # -------------------------------------------------
    # Update locality wise state automatically
    # -------------------------------------------------
    if locality_name:
        first_record = app_models.UserReg.objects.filter(
            locality__iexact=locality_name
        ).first()

        if first_record:
            state_name = (first_record.state or "").strip()

            app_models.LocalityWise.objects.update_or_create(
                locality=locality_name,
                defaults={"state": state_name}
            )

    # -------------------------------------------------
    # Apply filters
    # -------------------------------------------------
    filtered_queryset = apply_export_filters(
        app_models.UserReg.objects.order_by("sl_no"),
        selected_filter_rows,
    )

    export_headers, export_rows = get_userreg_export_dataset(
        selected_export_columns,
        filtered_queryset
    )

    # -------------------------------------------------
    # Locality export => fresh serial number
    # -------------------------------------------------
    if locality_name:
        export_headers.insert(0, "Sl No")
        export_rows = [
            [str(i)] + row
            for i, row in enumerate(export_rows, start=1)
        ]

    # -------------------------------------------------
    # Page size
    # -------------------------------------------------
    pdf_pagesize = landscape(legal) if locality_name else get_export_pdf_pagesize(len(export_headers))

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=pdf_pagesize,
        rightMargin=18,
        leftMargin=18,
        topMargin=16,
        bottomMargin=16,
        pageCompression=0,
    )

    styles = getSampleStyleSheet()

    # -------------------------------------------------
    # Styles
    # -------------------------------------------------
    title_style = styles["Title"].clone("title_style")
    title_style.fontSize = 17
    title_style.leading = 18
    title_style.spaceBefore = 0
    title_style.spaceAfter = 0

    contact_style = styles["BodyText"].clone("contact_style")
    contact_style.fontSize = 8
    contact_style.leading = 8.5
    contact_style.spaceBefore = 0
    contact_style.spaceAfter = 0

    report_title = "SOT Meeting Report"
    if locality_name:
        report_title += f" - {locality_name}"

    # -------------------------------------------------
    # Header
    # -------------------------------------------------
    header_data = [[
        Paragraph(f"<b>{report_title}</b>", title_style),

        Paragraph("""
        <b>Help Desk</b><br/>
        Bro. Jude - 9567413410<br/>
        Bro. Eldhose - 9497774409<br/>
        <b>Accommodation</b><br/>
        Bro. Edwin Sam - 9562324451<br/>
        Bro. Kiran Christo - 8086276237<br/>
        <b>Transportation</b><br/>
        Bro. Saju - 8281066319<br/>
        Bro. Sajo - 9061104753
        """, contact_style)
    ]]

    header_table = Table(header_data, colWidths=[560, 190])

    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    story = [
        header_table,
        Spacer(1, 0.06 * inch),
        Paragraph(f"Rows Included: {len(export_rows)}", styles["BodyText"]),
        Spacer(1, 0.08 * inch),
    ]

    # -------------------------------------------------
    # Data Table
    # -------------------------------------------------
    table_data = build_export_pdf_table_data(export_headers, export_rows, styles)

    available_width = pdf_pagesize[0] - document.leftMargin - document.rightMargin

    col_widths = get_export_pdf_col_widths(
        export_headers,
        export_rows,
        available_width
    )

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),

        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    # Right align money columns
    for idx, head in enumerate(export_headers):
        if head.lower() in ["total amount", "balance amount"]:
            table.setStyle(TableStyle([
                ("ALIGN", (idx, 1), (idx, -1), "RIGHT")
            ]))

    story.append(table)

    # -------------------------------------------------
    # Print only matched room / venue codes
    # -------------------------------------------------
    room_map = {
        "ANR": "Anna Residency",
        "AMR": "Amma Residency",
        "CFR": "Comfort Inn",
        "SS": "Silver Stone",
        "ARCADE": "Dreams Arcade",
        "GGH": "Government Guest House",
        "LUX ADB": "LUX Adobe",
        "LUX P": "LUX Palm Grove",
        "LUX S": "LUX Spicy",
        "LUX C": "LUX Castillo",
        "LUX AA": "LUX Apartment A",
        "LUX AB": "LUX Apartment B",
        "LUX AC": "LUX Apartment C",
        "LUX N": "LUX Nest",
        "LUX V": "LUX Villa",
        "SELF": "Self Accommodation",
    }

    venue_map = {
        "D1": "Dormitory 1",
        "D2": "Dormitory 2",
        "D3": "Dormitory 3",
        "D4": "Dormitory 4",
        "VC": "Venue Cubicle",
        "V DB": "Venue Double Bed",
    }

    used_rooms = (
        filtered_queryset
        .exclude(acc_on_1_2__isnull=True)
        .exclude(acc_on_1_2="")
        .values_list("acc_on_1_2", flat=True)
        .distinct()
    )

    found_room_codes = []
    found_venue_codes = []

    for item in used_rooms:
        txt = str(item).upper().strip()

        for code in room_map.keys():
            if txt == code or txt.startswith(code + " "):
                if code not in found_room_codes:
                    found_room_codes.append(code)

        for code in venue_map.keys():
            if txt == code or txt.startswith(code + " "):
                if code not in found_venue_codes:
                    found_venue_codes.append(code)

    # Accommodation
    if found_room_codes:
        story.append(Spacer(1, 0.15 * inch))
        story.append(Paragraph("<b>Accommodation Codes</b>", styles["Heading4"]))

        room_text = "<br/>".join(
            [f"{code} - {room_map[code]}" for code in found_room_codes]
        )

        story.append(Paragraph(room_text, styles["BodyText"]))

    # Venue
    if found_venue_codes:
        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph("<b>Venue Codes</b>", styles["Heading4"]))

        venue_text = "<br/>".join(
            [f"{code} - {venue_map[code]}" for code in found_venue_codes]
        )

        story.append(Paragraph(venue_text, styles["BodyText"]))

    # -------------------------------------------------
    # Build PDF
    # -------------------------------------------------
    document.build(story)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/pdf"
    )

    # -------------------------------------------------
    # Dynamic filename
    # -------------------------------------------------
    if locality_name and state_name:
        file_name = f"{locality_name}_{state_name}.pdf"
    elif locality_name:
        file_name = f"{locality_name}.pdf"
    else:
        file_name = "sot-meeting-report.pdf"

    file_name = file_name.replace(" ", "_")

    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    return response

def locality_register(request):
    context = build_common_context(request)
    locality_summary = None
    table=app_models.LocalityWise.objects.order_by("locality").values("locality", "state", "persons_count", "total_balance")

    if request.method == "POST":
        locality_frm = app_forms.LocalityRegisterForm(request.POST)
        if locality_frm.is_valid():
            locality = locality_frm.cleaned_data["locality"]
            locality_summary = get_locality_register_summary(locality)
            defaults = locality_frm.cleaned_data.copy()
            defaults["state"] = locality_summary["state"]
            defaults["persons_count"] = locality_summary["persons_count"]
            defaults["total_balance"] = Decimal(locality_summary["balance_amount_sum"])
            defaults.pop("locality", None)
            app_models.LocalityWise.objects.update_or_create(
                locality=locality,
                defaults=defaults,
            )
            return redirect("reg")
    else:
        locality_frm = app_forms.LocalityRegisterForm()

    context["locality_frm"] = locality_frm
    context["paid_localities"] = get_paid_locality_list()
    selected_locality = ""
    if locality_frm.is_bound:
        selected_locality = locality_frm.data.get("locality", "").strip()
    elif locality_frm.initial.get("locality"):
        selected_locality = str(locality_frm.initial.get("locality")).strip()

    if selected_locality:
        locality_summary = get_locality_register_summary(selected_locality)

    context["locality_register_summary"] = locality_summary
    return render(request, "account/regiser.html", context)


def get_locality_register_summary(locality):
    queryset = app_models.UserReg.objects.filter(locality__iexact=locality).order_by("sl_no")
    aggregates = queryset.aggregate(
        total_amount_sum=Sum("total_amount"),
        balance_amount_sum=Sum("balance_amount"),
    )
    first_record = queryset.first()
    return {
        "locality": locality,
        "state": first_record.state if first_record and first_record.state else "",
        "persons_count": queryset.count(),
        "brother_count": queryset.filter(bs__iexact="Brother").count(),
        "sister_count": queryset.filter(bs__iexact="Sister").count(),
        "children_count": queryset.filter(age__isnull=False, age__gte=5, age__lt=13).count(),
        "infant_count": queryset.filter(age__isnull=False, age__lt=5).count(),
        "teenager_count": queryset.filter(age__isnull=False, age__gte=13, age__lt=20).count(),
        "total_amount_sum": format_export_cell_value(aggregates.get("total_amount_sum") or Decimal("0")),
        "balance_amount_sum": format_export_cell_value(aggregates.get("balance_amount_sum") or Decimal("0")),
    }


def get_paid_locality_list():
    return list(
        app_models.LocalityWise.objects.exclude(payment_method="pending")
        .order_by("locality")
        .values("locality", "state", "total_paid", "payment_method")
    )


def get_search_paid_localities(query=""):
    queryset = app_models.LocalityWise.objects.exclude(payment_method="pending").order_by("locality")
    if query:
        queryset = queryset.filter(
            Q(locality__icontains=query)
            | Q(state__icontains=query)
            | Q(payment_method__icontains=query)
        )
    return list(queryset.values("locality", "state", "total_paid", "payment_method"))


def get_userreg_filter_options(selected_filters=None):
    selected_filters = selected_filters or {}
    options = []
    base_queryset = app_models.UserReg.objects.order_by("sl_no")
    for field_name in USERREG_FILTER_FIELDS:
        values = list(
            base_queryset.exclude(**{f"{field_name}__isnull": True})
            .exclude(**{field_name: ""})
            .order_by(field_name)
            .values_list(field_name, flat=True)
            .distinct()
        )
        options.append(
            {
                "name": field_name,
                "label": format_field_label(field_name),
                "values": [
                    {
                        "value": value,
                        "selected": selected_filters.get(field_name) == value,
                    }
                    for value in values
                ],
            }
        )
    return options


def resolve_userreg_filters(request):
    selected_filters = {}
    for field_name in USERREG_FILTER_FIELDS:
        selected_value = request.GET.get(field_name, "").strip()
        if selected_value:
            selected_filters[field_name] = selected_value
    return selected_filters


def apply_userreg_filters(queryset, selected_filters):
    filtered_queryset = queryset
    for field_name, selected_value in selected_filters.items():
        filtered_queryset = filtered_queryset.filter(**{field_name: selected_value})
    return filtered_queryset


def locality_register_summary(request):
    locality = request.GET.get("locality", "").strip()
    if not locality:
        return JsonResponse({"error": "Locality is required."}, status=400)
    return JsonResponse(get_locality_register_summary(locality))


def query(request):
    context = build_common_context(request)

    result = []
    columns = []
    error = ""
    query_text = ""

    if request.method == "POST":
        query_text = request.POST.get("query", "").strip()

        try:
            # allow only SELECT query for safety
            if not query_text.lower().startswith("select"):
                error = "Only SELECT query allowed."
            else:
                with connection.cursor() as cursor:
                    cursor.execute(query_text)

                    columns = [col[0] for col in cursor.description]
                    result = cursor.fetchall()

        except Exception as e:
            error = str(e)

    context.update({
        "query_text": query_text,
        "columns": columns,
        "result": result,
        "error": error,
    })

    return render(request, "account/query.html", context)

# views.py

def query_export_pdf(request):
    from io import BytesIO
    from django.http import HttpResponse
    from django.db import connection

    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        LongTable,
    )

    from reportlab.lib.pagesizes import landscape, legal
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    query_text = request.GET.get("query", "").strip()

    # --------------------------------------------
    # Allow only SELECT
    # --------------------------------------------
    if not query_text.lower().startswith("select"):
        return HttpResponse("Only SELECT query allowed")

    # --------------------------------------------
    # Run Query
    # --------------------------------------------
    try:
        with connection.cursor() as cursor:
            cursor.execute(query_text)

            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

    except Exception as e:
        return HttpResponse(str(e))

    # --------------------------------------------
    # PDF Setup
    # --------------------------------------------
    output = BytesIO()

    pagesize = landscape(legal)

    doc = SimpleDocTemplate(
        output,
        pagesize=pagesize,
        leftMargin=15,
        rightMargin=15,
        topMargin=15,
        bottomMargin=15,
    )

    styles = getSampleStyleSheet()
    story = []

    page_width = pagesize[0]
    usable_width = page_width - doc.leftMargin - doc.rightMargin

    # --------------------------------------------
    # Heading
    # --------------------------------------------
    story.append(
        Paragraph("<b>SQL Query Report</b>", styles["Title"])
    )
    story.append(Spacer(1, 0.10 * inch))

    story.append(
        Paragraph(f"<b>Total Rows:</b> {len(rows)}", styles["BodyText"])
    )
    story.append(Spacer(1, 0.08 * inch))

    # --------------------------------------------
    # Dynamic Table Data
    # --------------------------------------------
    table_data = []

    headers = ["Sl No"] + columns
    table_data.append(headers)

    for i, row in enumerate(rows, start=1):
        clean_row = [str(i)]

        for val in row:
            if val is None:
                clean_row.append("")
            else:
                clean_row.append(str(val))

        table_data.append(clean_row)

    # --------------------------------------------
    # Dynamic Column Width Control
    # overflow / underflow fit
    # --------------------------------------------
    col_count = len(headers)

    raw_widths = []

    for col_index in range(col_count):
        sample = []

        for row in table_data[:30]:
            if col_index < len(row):
                sample.append(str(row[col_index]))

        max_len = max([len(x) for x in sample], default=8)

        # minimum and max dynamic size
        width = max(45, min(max_len * 6, 180))

        raw_widths.append(width)

    total_width = sum(raw_widths)

    # if overflow => shrink all
    if total_width > usable_width:
        scale = usable_width / total_width
        col_widths = [w * scale for w in raw_widths]

    # if underflow => stretch equally
    elif total_width < usable_width:
        extra = (usable_width - total_width) / col_count
        col_widths = [w + extra for w in raw_widths]

    else:
        col_widths = raw_widths

    # --------------------------------------------
    # LongTable auto multipage
    # --------------------------------------------
    table = LongTable(
        table_data,
        colWidths=col_widths,
        repeatRows=1
    )

    table.setStyle(TableStyle([

        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),

        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),

        ("FONTSIZE", (0, 0), (-1, -1), 7),

        ("VALIGN", (0, 0), (-1, -1), "TOP"),

        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),

        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),

    ]))

    story.append(table)

    # --------------------------------------------
    # Build PDF
    # --------------------------------------------
    doc.build(story)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="sql_query_report.pdf"'
    )

    return response

# views.py
def idcard(request):
    from io import BytesIO
    from django.http import HttpResponse
    from django.shortcuts import render

    from reportlab.platypus import (
        BaseDocTemplate,
        PageTemplate,
        Frame,
        Paragraph,
        Table,
        TableStyle,
        Image,
        Spacer,
        KeepTogether,
    )

    from reportlab.lib.units import inch
    from reportlab.lib.pagesizes import A4, A3, legal, letter, landscape, portrait
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    # ----------------------------------------------------
    # OPEN PAGE
    # ----------------------------------------------------
    if request.method != "POST":
        context = build_common_context(request)
        return render(request, "account/idcard.html", context)

    # ----------------------------------------------------
    # DEFAULT CARD SIZE
    # ----------------------------------------------------
    CARD_W = 3.37 * inch
    CARD_H = 2.125 * inch

    # ----------------------------------------------------
    # PAGE OPTIONS
    # ----------------------------------------------------
    paper = request.POST.get("paper_size", "A4").upper()
    orient = request.POST.get("orientation", "portrait").lower()

    page_map = {
        "A4": A4,
        "A3": A3,
        "LEGAL": legal,
        "LETTER": letter,
    }

    page_size = page_map.get(paper, A4)

    if orient == "landscape":
        page_size = landscape(page_size)
    else:
        page_size = portrait(page_size)

    PAGE_W, PAGE_H = page_size

    # ----------------------------------------------------
    # GRID CALCULATION
    # ----------------------------------------------------
    margin = 0.18 * inch
    gap_x = 0.08 * inch
    gap_y = 0.08 * inch

    usable_w = PAGE_W - (margin * 2)
    usable_h = PAGE_H - (margin * 2)

    cols = int((usable_w + gap_x) // (CARD_W + gap_x))
    rows = int((usable_h + gap_y) // (CARD_H + gap_y))

    cols = max(cols, 1)
    rows = max(rows, 1)

    grid_w = cols * CARD_W + (cols - 1) * gap_x
    grid_h = rows * CARD_H + (rows - 1) * gap_y

    start_x = (PAGE_W - grid_w) / 2
    start_y = PAGE_H - ((PAGE_H - grid_h) / 2)

    # ----------------------------------------------------
    # FRAMES
    # ----------------------------------------------------
    frames = []

    for r in range(rows):
        for c in range(cols):

            x = start_x + c * (CARD_W + gap_x)
            y = start_y - ((r + 1) * CARD_H) - (r * gap_y)

            frames.append(
                Frame(
                    x, y,
                    CARD_W,
                    CARD_H,
                    leftPadding=2,
                    rightPadding=2,
                    topPadding=2,
                    bottomPadding=2,
                    showBoundary=1,
                )
            )

    # ----------------------------------------------------
    # PDF
    # ----------------------------------------------------
    output = BytesIO()

    doc = BaseDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=0,
        rightMargin=0,
        topMargin=0,
        bottomMargin=0,
    )

    doc.addPageTemplates([
        PageTemplate(id="cards", frames=frames)
    ])

    styles = getSampleStyleSheet()

    head_style = ParagraphStyle(
        "head",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        alignment=1,
        leading=10,
        textColor=colors.white,
    )

    label_style = ParagraphStyle(
        "label",
        parent=styles["Normal"],
        fontSize=5,
        leading=5.4,
        fontName="Helvetica",
    )

    # ----------------------------------------------------
    # AUTO TEXT FIX
    # ----------------------------------------------------
    def trim_text(txt, max_len=23):
        txt = str(txt or "").strip()
        if len(txt) <= max_len:
            return txt
        return txt[:max_len] + "..."

    def font_size(txt):
        ln = len(str(txt))
        if ln > 28:
            return 4
        elif ln > 22:
            return 4.5
        elif ln > 16:
            return 5
        return 5.6

    # ----------------------------------------------------
    # IMAGES
    # ----------------------------------------------------
    logo_top = "static/images/image1.png"      # header image
    logo_middle = "static/images/image1.png"   # transparent middle image

    users = app_models.UserReg.objects.order_by("sl_no")

    story = []

    # ----------------------------------------------------
    # GENERATE BULK CARDS
    # ----------------------------------------------------
    for user in users:

        block = []

        # --------------------------------------------
        # HEADER IMAGE + TITLE
        # --------------------------------------------
        try:
            top_img = Image(
                logo_top,
                width=0.30 * inch,
                height=0.30 * inch
            )
        except:
            top_img = ""

        header = Table(
            [[top_img, Paragraph("SOT MEETING", head_style)]],
            colWidths=[0.35 * inch, CARD_W - 0.45 * inch],
            rowHeights=[0.30 * inch]
        )

        header.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.green),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ]))

        block.append(header)
        block.append(Spacer(1, 1))

        # --------------------------------------------
        # TRANSPARENT CENTER IMAGE
        # --------------------------------------------
        try:
            mid_img = Image(
                logo_middle,
                width=0.42 * inch,
                height=0.42 * inch,
                mask='auto'
            )

            img_tbl = Table(
                [[mid_img]],
                colWidths=[CARD_W - 6]
            )

            img_tbl.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))

            block.append(img_tbl)
            block.append(Spacer(1, 1))

        except:
            pass

        # --------------------------------------------
        # DETAILS
        # --------------------------------------------
        fields = [
            ["Name", user.name],
            ["Locality", user.locality],
            ["State", user.state],
            ["Lang", user.language],
            ["Room", user.acc_on_1_2],
            ["Bus", getattr(user, "bus_no", "")],
        ]

        rows_data = []

        for label, val in fields:

            clean = trim_text(val, 24)

            val_style = ParagraphStyle(
                "val",
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                fontSize=font_size(clean),
                leading=font_size(clean) + 0.4,
                wordWrap="CJK",
            )

            rows_data.append([
                Paragraph(label, label_style),
                Paragraph(clean, val_style)
            ])

        body = Table(
            rows_data,
            colWidths=[0.58 * inch, CARD_W - 0.76 * inch]
        )

        body.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        block.append(body)

        story.append(KeepTogether(block))

    # ----------------------------------------------------
    # BUILD PDF
    # ----------------------------------------------------
    doc.build(story)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/pdf"
    )

    response["Content-Disposition"] = 'attachment; filename="SOT_ID_CARDS.pdf"'

    return response